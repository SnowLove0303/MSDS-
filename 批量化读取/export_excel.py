# -*- coding: utf-8 -*-
"""MSDS 全量信息汇总 → Excel 宽表.

首行 = 字段标题, 首列 = 产品型号, 按产品大类(型号前缀)从上到下排序,
第二行起写数据. 保证所有信息检索完整.

从 MSDS数据库.sqlite 读取:
  - msds_documents  主表 (S0 页眉 + 元数据)
  - msds_fields     EAV (S1-S16 全节)
  - msds_components 成分 1:N

列设计 (约 200 列):
  - 首列 产品型号; 次列 产品大类; 再文件名称/语言/厂家/版本/修订日期/缺失节
  - 标准字段列 (field_key 命中 _NORM 映射) 单独成列, 列头 S{n}·字段名
  - 高频保留字段列 (未归一化但覆盖 >= 15 文档) 单独成列, 保证常用信息完整
  - 节完整信息列: 吸收该节其余未分列字段, 保证信息零丢失
  - 成分分列: 成分1..N 名称/CAS/含量 (最大成分数定列)

用法:
  python export_excel.py [数据库.sqlite] [-o 输出.xlsx]
"""
from __future__ import annotations

import re
import sqlite3
import sys
import collections
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_db import _NORM

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("需要 openpyxl: pip install openpyxl")

_STDSET = set(_NORM.keys())
_MAX_COMP = 6          # 全库实测最大成分数
_MIN_RAW_DOCS = 15     # 未归一化字段单独成列的覆盖阈值
_N_AUX = 9             # 辅助列数 (含 S3·成分原文)


def model_from_filename(fn: str) -> str:
    """从文件名提取型号代码 (product_name 为空时兜底)."""
    fn = fn.replace(".docx", "")
    m = re.match(r"^([A-Z]{1,4}-\d+[A-Za-z0-9]*)", fn)
    if m:
        return re.sub(r"(?i)(?:msds|ms|—|-)\s*$", "", m.group(1)).strip()
    m = re.match(r"^([A-Za-z]{1,4}\d+[A-Za-z0-9]*)", fn)
    if m:
        return m.group(1)
    return fn.strip()


def category_of(model: str) -> str:
    """产品大类 = 型号前缀字母 (无型号归'其他')."""
    m = re.match(r"^([A-Z]{1,4})-\d", model or "")
    return m.group(1) if m else "其他"


def nat_key(model: str):
    """自然排序键: 数字段按数值, 字母段按原文."""
    parts = re.split(r"(\d+)", model or "")
    return [int(p) if p.isdigit() else p for p in parts]


def load_all(db: Path):
    conn = sqlite3.connect(str(db))
    conn.row_factory = sqlite3.Row

    # 主表
    docs = [dict(r) for r in conn.execute(
        "SELECT * FROM msds_documents ORDER BY id")]
    # 字段表 (doc_id -> list[(section, field_key, value)])
    fields = collections.defaultdict(list)
    for r in conn.execute(
            "SELECT doc_id, section, field_key, value FROM msds_fields"):
        fields[r["doc_id"]].append((r["section"], r["field_key"], r["value"] or ""))
    # 字段表各 (section, field_key) 覆盖文档数 → 决定列
    col_docs = collections.Counter()
    for r in conn.execute(
            "SELECT section, field_key, COUNT(DISTINCT doc_id) c "
            "FROM msds_fields GROUP BY 1,2"):
        col_docs[(r["section"], r["field_key"])] = r["c"]
    # 成分
    comps = collections.defaultdict(list)
    for r in conn.execute(
            "SELECT doc_id, comp_idx, comp_name, cas, conc "
            "FROM msds_components ORDER BY doc_id, comp_idx"):
        comps[r["doc_id"]].append((r["comp_name"] or "", r["cas"] or "",
                                   r["conc"] or ""))
    conn.close()
    return docs, fields, col_docs, comps


def build_columns(col_docs):
    """返回 (columns, col_sections) columns: [(key, header), ...]"""
    # 分组
    std_items = []   # (section, field_key, docs)
    raw_items = []
    sec_low = collections.defaultdict(list)   # 节 -> [(field_key, docs)]
    for (sec, fk), c in col_docs.items():
        if not fk or not fk.strip():
            # 无标签正文 (note): 并入该节"完整信息"列, 保证不丢
            sec_low[sec].append(("", c))
            continue
        if fk in _STDSET:
            std_items.append((sec, fk, c))
        else:
            raw_items.append((sec, fk, c))
    std_items.sort(key=lambda x: (-x[2], x[0], x[1]))
    raw_items.sort(key=lambda x: (-x[2], x[0], x[1]))

    # 高频保留原样字段单独成列
    high_raw = [(s, f, c) for s, f, c in raw_items if c >= _MIN_RAW_DOCS]
    # 低频保留原样字段 → 完整信息列
    for s, f, c in raw_items:
        if c < _MIN_RAW_DOCS:
            sec_low[s].append((f, c))

    columns = []   # (key, header)
    # 辅助列 (不占 field 值)
    columns += [("__model", "产品型号"), ("__cat", "产品大类"),
                ("__file", "文件名称"), ("__lang", "语言"),
                ("__vendor", "厂家"), ("__ver", "版本"),
                ("__rev", "修订日期"), ("__missing", "缺失节"),
                ("__s3raw", "S3·成分原文")]
    # S0 之后的字段列, 按节 0..16 排序 (S0 无字段列)
    by_sec = collections.defaultdict(list)
    for sec, fk, c in std_items:
        by_sec[sec].append((fk, c))
    for sec, fk, c in high_raw:
        by_sec[sec].append((fk, c))
    for sec in sorted(by_sec):
        items = sorted(by_sec[sec], key=lambda x: (-x[1], x[0]))
        for fk, c in items:
            columns.append((("F", sec, fk), f"S{sec}·{fk}"))
        if sec_low.get(sec):
            columns.append((("Z", sec, None), f"S{sec}·完整信息"))
    # 成分列
    for i in range(1, _MAX_COMP + 1):
        columns.append((("C", i, "name"), f"S3·成分{i}名称"))
        columns.append((("C", i, "cas"), f"S3·成分{i}CAS"))
        columns.append((("C", i, "conc"), f"S3·成分{i}含量"))
    return columns, sec_low


def main(argv=None):
    argv = list(argv if argv is not None else sys.argv[1:])
    out = None
    db = None
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("-o", "--out") and i + 1 < len(argv):
            out = argv[i + 1]; i += 2
        elif a in ("-h", "--help"):
            print(__doc__); return 0
        else:
            db = a; i += 1
    if not db:
        db = r"F:\正式项目与模块化内容\Word 覆写模块\数据库\MSDS数据库.sqlite"
    if not out:
        out = str(Path(db).with_suffix("")) + "_全量汇总.xlsx"
    db = Path(db); out = Path(out)

    docs, fields, col_docs, comps = load_all(db)
    columns, sec_low = build_columns(col_docs)

    # 组装行 (按产品大类 + 型号自然排序)
    rows = []
    for d in docs:
        model = (d.get("product_name") or "").strip()
        if not model:
            model = model_from_filename(d.get("file_name") or "")
        cat = category_of(model)
        row = {"__model": model, "__cat": cat,
               "__file": d.get("file_name", ""), "__lang": d.get("language", ""),
               "__vendor": d.get("template_vendor", ""),
               "__ver": d.get("version", "") or "", "__rev": d.get("revision_date", "") or "",
               "__missing": d.get("missing_sections", "") or "",
               "__s3raw": d.get("s3_原文", "") or ""}
        # 字段值
        fmap = collections.defaultdict(list)   # (sec, fk) -> [value]
        for sec, fk, val in fields.get(d["id"], []):
            if val:
                fmap[(sec, fk)].append(val)
        # 完整信息列内容: 每节未分列的保留原样字段
        for sec, lows in sec_low.items():
            lines = []
            for fk, c in lows:
                vals = fmap.get((sec, fk))
                if vals:
                    txt = "; ".join(vals)
                    lines.append(f"{fk}: {txt}" if fk else txt)
            row[("Z", sec, None)] = "\n".join(lines)
        # 标准/高频字段值
        for key, header in columns:
            if not isinstance(key, tuple):
                continue
            kind = key[0]
            if kind == "F":
                sec, fk = key[1], key[2]
                vals = fmap.get((sec, fk))
                if vals:
                    row[key] = "\n".join(vals)
            elif kind == "C":
                _, idx, part = key
                cl = comps.get(d["id"], [])
                if idx - 1 < len(cl):
                    comp = cl[idx - 1]
                    row[key] = {"name": comp[0], "cas": comp[1],
                                "conc": comp[2]}[part]
        rows.append(row)

    # 排序: 产品大类字母序 (无型号"其他"放最后), 类内按型号自然排序
    rows.sort(key=lambda r: (r["__cat"] == "其他", r["__cat"], nat_key(r["__model"])))

    # 写 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "MSDS全量汇总"
    headers = [h for _, h in columns]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    for ridx, r in enumerate(rows, 2):
        for cidx, (key, h) in enumerate(columns, 1):
            v = r.get(key)
            cell = ws.cell(row=ridx, column=cidx, value=v if v not in (None, "") else "")
            cell.alignment = wrap
            cell.border = border
            # 节分组表头: 非辅助列按 (sec,kind) 着色区分
        # 大类分组行内交替底纹 (浅灰)
        if r["__cat"] != "其他" and rows.index(r) % 2 == 0:
            pass

    # 大类分隔: 整行浅灰底纹 (可选, 简单按奇偶)
    # 列宽
    from openpyxl.utils import get_column_letter
    for idx, (key, h) in enumerate(columns, 1):
        if h == "S3·成分原文":
            width = 50
        elif idx <= _N_AUX:
            width = 14
        else:
            width = 18
        if h.startswith("S3·成分") and "CAS" in h:
            width = 14
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows)+1}"

    # 大类分组: 每大类首行整行浅蓝底纹 + 型号列加粗, 从上到下清晰分组
    cat_fill = PatternFill("solid", fgColor="DCE6F1")
    prev_cat = None
    for ridx, r in enumerate(rows, 2):
        cat = r["__cat"]
        if cat == prev_cat:
            continue
        prev_cat = cat
        for cidx in range(1, len(columns) + 1):
            ws.cell(row=ridx, column=cidx).fill = cat_fill
        ws.cell(row=ridx, column=1).font = Font(bold=True)

    wb.save(str(out))
    print(f"导出完成: {len(rows)} 行 × {len(columns)} 列 -> {out}")
    print(f"列构成: 辅助 {_N_AUX} / 字段列 {len(columns)-_N_AUX-_MAX_COMP*3} / 成分 {_MAX_COMP*3}")
    # 分类统计
    cats = collections.Counter(r["__cat"] for r in rows)
    print("产品大类分布:", dict(sorted(cats.items())))
    return 0


if __name__ == "__main__":
    sys.exit(main())