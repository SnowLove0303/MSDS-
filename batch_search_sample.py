# -*- coding: utf-8 -*-
"""抽样 MSDS 批量检索 + 写 Excel (验证大批量链路).

用法:
  python batch_search_sample.py --n 10 --seed 42 --out 输出目录
                                [--search 词1 词2 ...]

流程:
  1. 从 MSDS 库 646 文件中按 seed 抽 n 个 (跨中/英、冠志/国彩)
  2. 逐个 read_msds → 内存总库 add_product (完整入库)
  3. 多关键词批量检索 (search_products: 标签/内容/成分/型号 双轨命中)
  4. 写一个 Excel 三页: 宽表 | 检索命中 | 抽样清单
"""
from __future__ import annotations

import argparse
import random
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from core.docx_reader import read_msds
from db import catalog
from db.schema import connect
from db.export import export_wide

MSDS_DIR = Path(r"F:\正式项目与模块化内容\Word 覆写模块\数据库\MSDS")

# 能力矩阵: 中文标准字段 / 中文内容 / 英文 raw_label 双轨 / S9 值 / 成分CAS / 成分名
DEFAULT_SEARCHES = ["毒性", "供应商", "supplier", "闪点",
                    "28182-81-2", "polyether"]


def pick_sample(n: int, seed: int) -> list[Path]:
    """固定 seed 抽样: 中/英 × 冠志/国彩 四象限均摊 + 随机补足."""
    rng = random.Random(seed)
    all_docs = sorted(MSDS_DIR.rglob("*.docx"))
    if n >= len(all_docs):
        return all_docs
    dirs = {}
    for d in all_docs:
        rel = d.parent.relative_to(MSDS_DIR)
        key = str(rel)
        dirs.setdefault(key, []).append(d)
    picked: list[Path] = []
    # 每象限均摊一份, 余数随机补足
    keys = sorted(dirs)
    per = n // len(keys)
    for k in keys:
        picked += rng.sample(dirs[k], min(per, len(dirs[k])))
    while len(picked) < n:
        cand = rng.choice(all_docs)
        if cand not in picked:
            picked.append(cand)
    return sorted(picked, key=lambda p: p.name)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--n", type=int, default=10)
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--out", default=str(MSDS_DIR.parent / "_抽样检索"),
                    help="Excel 输出目录")
    ap.add_argument("--search", nargs="*", default=DEFAULT_SEARCHES)
    args = ap.parse_args(argv)

    docs = pick_sample(args.n, args.seed)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    conn = connect(":memory:")
    samples = []
    for d in docs:
        try:
            result = read_msds(d)
            cat = d.parent.parent.name + "-" + d.parent.name  # 中文-冠志 等
            pid, status = catalog.add_product(conn, result, category=cat)
            samples.append({
                "pid": pid, "file": d.name, "rel": str(d.relative_to(MSDS_DIR)),
                "model": catalog.extract_model_name(d.name), "category": cat,
                "status": status, "fields": len(catalog.get_fields(conn, pid)),
                "components": len(catalog.get_components(conn, pid)),
            })
        except Exception as exc:
            print(f"⚠️ 读取失败 {d.name}: {exc}")
    if not samples:
        print("无成功入库样本")
        return 1

    # ---- 批量检索 ----
    print(f"\n=== 抽样 {len(docs)} 个 → 入库 {len(samples)} 个 ===")
    hits_rows = []
    for q in args.search:
        res = catalog.search_products(conn, q)
        models = [r["model_name"] for r in res]
        print(f"搜「{q}」→ 命中 {len(models)}: {', '.join(models) or '无'}")
        hits_rows.append({"query": q, "count": len(models),
                          "models": "、".join(models)})

    # ---- 写 Excel ----
    xlsx = out_dir / f"抽样检索_n{args.n}_seed{args.seed}.xlsx"
    stats = export_wide(conn, xlsx)
    print(f"✅ 宽表 → {xlsx}")

    # 追加检索命中 sheet + 抽样清单 sheet
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb.create_sheet("检索命中")
    ws.append(["查询词", "命中数", "命中型号"])
    for r in hits_rows:
        ws.append([r["query"], r["count"], r["models"]])
    for ci, w in enumerate((18, 8, 90), 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws2 = wb.create_sheet("抽样清单")
    ws2.append(["型号", "语言-厂商", "入库状态", "字段数", "成分数", "文件"])
    for s in samples:
        ws2.append([s["model"], s["category"], s["status"],
                    s["fields"], s["components"], s["file"]])
    for ci, w in enumerate((16, 16, 12, 8, 8, 70), 1):
        ws2.column_dimensions[chr(64 + ci)].width = w
    wb.save(xlsx)
    print(f"✅ 检索命中+抽样清单 → {xlsx} ({len(hits_rows)} 查询 / {len(samples)} 型号)")

    import json
    (out_dir / "抽样检索_report.json").write_text(
        json.dumps({"docs": len(docs), "inserted": len(samples),
                    "hits": hits_rows, "samples": samples},
                   ensure_ascii=False, indent=2), encoding="utf-8")
    return 0


if __name__ == "__main__":
    sys.exit(main())
