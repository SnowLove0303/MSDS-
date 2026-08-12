# -*- coding: utf-8 -*-
"""生成 MSDS 入库总表 (透视结构). 供 GUI「导出 Excel 库表」与命令行 tools 复用.

透视结构:
  - 第一行: Section 节标题 (合并单元格, 每节跨其全部列)
  - 第二行: 该节下的小标题/标签 (sub 小标题 + field 字段标签, 逐列)
  - 第三行起: 每型号一行, 按 节→标签 一一对照填入内容
  - A1 / A2 留空, A 列为型号
  - 成分表 (S3) 展开为 名称/CAS/含量 每成分三列

围栏 (异常防护):
  - 自动过滤 Word 临时文件 (~$ 开头)
  - 单个文件读取失败 → 跳过并记录失败清单, 不中断整体导出
  - 目录无 docx / 全部失败 → 明确报错

命令行用法:
  python tools/build_pivot_table.py <入库目录> <输出xlsx>
"""
from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .docx_reader import read_msds

# ---- 样式 ----
FONT = "Microsoft YaHei"
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")   # 深蓝: 节标题行
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")    # 浅蓝: sub 小标题
TAG_FILL = PatternFill("solid", fgColor="EDEDED")    # 浅灰: 字段标签
MODEL_FILL = PatternFill("solid", fgColor="F2F2F2")  # 型号列
NOTE_FILL = PatternFill("solid", fgColor="FFF7E6")   # 总结句
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _model_of(result, file_name: str) -> str:
    """型号: 优先取 S1 产品名称, 回退文件名前缀."""
    sec = result.sections.get(1)
    if sec:
        for row in sec.iter_rows():
            if row.kind == "field" and row.label == "产品名称" and row.value.strip():
                return row.value.strip()
    return re.split(r"\s+", Path(file_name).stem)[0]


def _collect_columns(results, max_comp: int) -> list[tuple[int, list[tuple[str, str]]]]:
    """收集每节列组 (保持首见顺序): (节号, [(kind, label), ...]).

    kind: sub / field / note / comp
    """
    sec_order: list[int] = []
    sec_cols: dict[int, list[tuple[str, str]]] = {}
    seen: set[tuple[int, str, str]] = set()

    def add(num: int, kind: str, label: str):
        if (num, kind, label) in seen:
            return
        seen.add((num, kind, label))
        if num not in sec_order:
            sec_order.append(num)
            sec_cols.setdefault(num, [])
        sec_cols[num].append((kind, label))

    for r in results:
        for num in sorted(r.sections):
            sec = r.sections[num]
            for row in sec.iter_rows():
                if row.kind == "section":
                    continue
                if row.kind == "sub":
                    add(num, "sub", row.label or row.seq or "(标题)")
                elif row.kind == "field":
                    if row.span and not row.label.strip():
                        add(num, "note", "(总结句)")
                    else:
                        add(num, "field", row.label)
                elif row.kind == "note":
                    add(num, "note", "(总结句)")
            # 成分表 → 展开为每成分三列
            if sec.is_component_table:
                for i in range(max_comp):
                    add(num, "comp", f"成分{i+1}名称")
                    add(num, "comp", f"成分{i+1}CAS")
                    add(num, "comp", f"成分{i+1}含量")
    return [(n, sec_cols[n]) for n in sorted(sec_order)]


def _value_of(result, num: int, kind: str, label: str) -> str:
    sec = result.sections.get(num)
    if not sec:
        return ""
    if kind == "sub":
        return ""                      # 小标题自身无内容
    if kind == "note":
        # 该节全部总结句合并 (保持原顺序, 换行连接)
        vals = [row.value for row in sec.iter_rows()
                if (row.kind in ("note", "field")) and row.span and not row.label.strip()]
        return "\n".join(v for v in vals if v.strip())
    if kind == "comp":
        m = re.match(r"成分(\d+)(名称|CAS|含量)", label)
        if not m:
            return ""
        idx = int(m.group(1)) - 1
        if idx >= len(sec.components):
            return ""
        c = sec.components[idx]
        return {"名称": c.name, "CAS": c.cas, "含量": c.conc}[m.group(2)]
    # field: 精确标签匹配
    for row in sec.iter_rows():
        if row.kind == "field" and row.label == label:
            return row.value
    return ""


def _clean(v: str) -> str:
    return v.replace("\r\n", "\n").strip()


def build_pivot_table(in_dir: str | Path, out_path: str | Path) -> dict:
    """生成透视总表.

    Returns:
        {files, cols, rows, sections, failed, out}
        files: 成功读取并入库的文件数; failed: 读取失败文件名清单
    """
    in_dir = Path(in_dir)
    files = sorted(p for p in in_dir.glob("*.docx")
                   if not p.name.startswith("~$"))
    if not files:
        raise FileNotFoundError(f"目录中未找到 docx 文件: {in_dir}")

    # 围栏: 单个文件读取失败 → 跳过并记录, 不中断整体导出
    ok_files: list[Path] = []
    results = []
    failed: list[str] = []
    for p in files:
        try:
            results.append(read_msds(p))
            ok_files.append(p)
        except Exception as exc:
            failed.append(f"{p.name} ({exc.__class__.__name__})")
    if not ok_files:
        raise RuntimeError(f"目录下 {len(files)} 个 docx 全部读取失败:\n" + "\n".join(failed))

    max_comp = max((len(s.components)
                    for r in results for s in r.sections.values()
                    if s.is_component_table), default=0)
    columns = _collect_columns(results, max_comp)

    wb = Workbook()
    ws = wb.active
    ws.title = "入库总表"

    # ---- 表头两行 ----
    col = 2  # A 列留作型号
    span_ranges: list[tuple[int, int, str]] = []   # (start_col, end_col, 节标题)
    for num, cols in columns:
        start = col
        for kind, label in cols:
            c = ws.cell(2, col, label)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = SUB_FILL if kind == "sub" else TAG_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            col += 1
        end = col - 1
        span_ranges.append((start, end, num))

    # ---- 第一行: 节标题 ----
    for start, end, num in span_ranges:
        sec = results[0].sections.get(num)
        title = sec.full_title if sec else f"第{num}节"
        if num == 0:
            title = "0 页眉/页脚"
        ws.merge_cells(start_row=1, start_column=start,
                       end_row=1, end_column=end)
        c = ws.cell(1, start, title)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        for cc in range(start, end + 1):
            ws.cell(1, cc).border = BORDER
        ws.row_dimensions[1].height = 28

    # ---- A1 / A2 留空 ----
    ws.cell(1, 1, "")
    ws.cell(2, 1, "")
    for r_ in (1, 2):
        c = ws.cell(r_, 1)
        c.fill = HEAD_FILL if r_ == 1 else TAG_FILL

    # ---- 数据行 ----
    for ri, (p, r) in enumerate(zip(ok_files, results), start=3):
        model = _model_of(r, p.name)
        mc = ws.cell(ri, 1, model)
        mc.font = Font(name=FONT, size=10, bold=True)
        mc.fill = MODEL_FILL
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = BORDER

        col = 2
        for num, cols in columns:
            for kind, label in cols:
                v = _clean(_value_of(r, num, kind, label))
                c = ws.cell(ri, col, v)
                c.font = Font(name=FONT, size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = BORDER
                if kind == "note" and v:
                    c.fill = NOTE_FILL
                col += 1
        ws.row_dimensions[ri].height = 60

    # ---- 列宽 / 冻结 ----
    ws.column_dimensions["A"].width = 20
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "B3"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "files": len(ok_files),
        "cols": col - 1,
        "rows": len(ok_files),
        "sections": len(columns),
        "failed": failed,
        "out": str(out_path),
    }
