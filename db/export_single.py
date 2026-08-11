# -*- coding: utf-8 -*-
"""单份 MSDS → 数据库格式 Excel (双 Sheet).

Sheet1 数据库宽表: 行=1, 列 = 基础元数据 + 非空标准字段列 + 说明列 + 成分列.
  - 数据列严格按节 1→16 升序 + 节内首现行序
  - 非空 field 与 note 均入列 (纯 note 节如 S16 不再整节消失)
  - S3 成分列插在 S3 节内 (紧跟 S3 字段之后, 非所有节末尾)
  - 空值字段列跳过 (消除大量空列混乱)

Sheet2 完整明细: 逐行全量 (禁止缺失), 九列结构:
  节 | 行类型 | 序号 | 原始标签 | 标准字段 | 值 | 父级 | 可编辑 | 备注
  - 序号列: 保留原文序号 (9.1~9.19), 恢复原始顺序与上下级
  - 父级列: 节→子标题→字段 父子链; 孤儿节点(无标签)挂回最近字段
  - 备注列: 分组标题/空值/页眉页脚 标注移出值列, 值列只放真实内容
  - 成分拆行: 每成分拆 名称/CAS/含量 三行, 与宽表成分列一一对应
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .catalog import resolve_field_key


def export_single(result, conn, out_path, *, category="", model_name=None) -> dict:
    """生成单文件数据库格式 Excel. 返回 {'wide_cols','detail_rows','components','wide_keys'}."""
    from .catalog import extract_model_name
    model = model_name or extract_model_name(result.file_name)

    # 产品元数据 (从 S0 页眉字段提取)
    prod_name = ver = rev = ""
    sec0 = result.sections.get(0)
    if sec0:
        for row in sec0.iter_rows():
            lab = (row.label or "").lower()
            if not row.value:
                continue
            if "产品名称" in lab or "product" in lab:
                prod_name = prod_name or row.value
            elif "version" in lab:
                ver = ver or row.value
            elif "修订" in lab or "revision" in lab:
                rev = rev or row.value

    # 构建宽表列: 严格节序(1→16) + 节内首现行序.
    # 非空 field 与 note 均入列 (纯 note 节如 S16 不再整节消失);
    # S3 成分列插在产品类型之后 (S3 节内), 而非排到所有节字段之后.
    columns: list = []      # (节, 显示key, 类型) 类型 ∈ field / note / comp
    col_pos: dict = {}      # (节, key) → columns 索引
    vals: dict = {}         # (节, key) → 值 (同 key 多行 field 合并去重)
    note_seq: dict = {}
    for n in range(1, 17):
        sec = result.sections.get(n)
        if not sec:
            continue
        note_seq[n] = 0
        for row in sec.iter_rows():
            v = (row.value or "").strip()
            if row.kind == "field":
                k = resolve_field_key(conn, n, row.label) if row.label else ""
                if not k or not v:
                    continue
                tup = (n, k)
                if tup not in col_pos:
                    col_pos[tup] = len(columns)
                    columns.append((n, k, "field"))
                if tup in vals:
                    for line in v.split("\n"):
                        line = line.strip()
                        if line and line not in vals[tup].split("\n"):
                            vals[tup] = (vals[tup] + "\n" + line).strip() if vals[tup] else line
                else:
                    vals[tup] = v
            elif row.kind == "note" and v:
                note_seq[n] += 1
                key = f"说明{note_seq[n]}"
                tup = (n, key)
                col_pos[tup] = len(columns)
                columns.append((n, key, "note"))
                vals[tup] = v
        if n == 3 and sec.components:
            for i in range(1, len(sec.components) + 1):
                for part in ("名称", "CAS", "含量"):
                    columns.append((3, f"成分{i}{part}", "comp"))

    sec3 = result.sections.get(3)
    comps = list(sec3.components) if sec3 and sec3.components else []

    # ================= Sheet1 宽表 =================
    wb = Workbook()
    ws = wb.active
    ws.title = "数据库宽表"
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="2F5597")
    hdr_font = Font(name="Microsoft YaHei", size=9, bold=True, color="FFFFFF")
    val_font = Font(name="Microsoft YaHei", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(name="Microsoft YaHei", size=9, bold=True)

    base_headers = ["型号", "一级类目", "产品名称", "版本", "修订日期", "页眉", "页脚"]
    headers = base_headers + [f"S{n}.{k}" for n, k, _t in columns]

    def set_hdr(cell, h):
        cell.value = h
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap
        cell.border = border

    for ci, h in enumerate(headers, 1):
        set_hdr(ws.cell(1, ci), h)
    ws.freeze_panes = "B2"

    def cell(ri, ci, v, font=None):
        c = ws.cell(ri, ci, v)
        c.font = font or val_font
        c.alignment = wrap
        c.border = border
        return c

    ri = 2
    col_of = {}
    for ci, (n, k, _t) in enumerate(columns, start=len(base_headers) + 1):
        col_of[(n, k)] = ci
    cell(ri, 1, model, bold)
    cell(ri, 2, category)
    cell(ri, 3, prod_name)
    cell(ri, 4, ver)
    cell(ri, 5, rev)
    cell(ri, 6, result.header)
    cell(ri, 7, result.footer)
    for tup, v in vals.items():
        if tup in col_of:
            cell(ri, col_of[tup], v)
    # 成分列: 已按 S3 节序插入 columns, 按映射写入 (不再排到所有节之后)
    for i, comp in enumerate(comps, 1):
        for j, part in enumerate(("名称", "CAS", "含量")):
            cell(ri, col_of[(3, f"成分{i}{part}")], (comp.name, comp.cas, comp.conc)[j])

    ws.column_dimensions["A"].width = 12
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 14 if c not in (6, 7) else 28
    for c in range(8, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ================= Sheet2 完整明细 =================
    # 九列: 节|行类型|序号|原始标签|标准字段|值|父级|可编辑|备注
    #  - 序号列: 保留原文序号 (row.seq) — 恢复原始顺序与上下级
    #  - 父级列: 节→子标题→字段 父子链 — 孤儿(无标签)挂回最近字段
    #  - 备注列: 分组标题/空值/页眉页脚 标注移出值列 — 值列只放真实内容
    #  - 成分拆行: 名称/CAS/含量 三行, 与宽表成分列一一对应
    ws2 = wb.create_sheet("完整明细")
    cols2 = ["节", "行类型", "序号", "原始标签", "标准字段", "值", "父级", "可编辑", "备注"]
    for ci, h in enumerate(cols2, 1):
        set_hdr(ws2.cell(1, ci), h)
    ws2.freeze_panes = "B2"

    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    sec_fill = PatternFill("solid", fgColor="2F5597")
    sec_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    sub_font = Font(name="Microsoft YaHei", size=9, bold=True)
    note_font = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")

    kind_cn = {"field": "字段", "sub": "子标题", "note": "说明",
               "component": "成分", "component_header": "成分表头"}
    rows_detail = 0
    r2 = 2

    def detail_row(sec_no, kind, seq, lab, std, val, parent, editable, note,
                   font, fill=None):
        nonlocal r2, rows_detail
        vals = [sec_no, kind, seq, lab, std, val, parent, editable, note]
        for cc, v in enumerate(vals, 1):
            x = ws2.cell(r2, cc, v)
            x.font = font
            x.alignment = wrap
            x.border = border
            if fill:
                x.fill = fill
        r2 += 1
        rows_detail += 1

    for n in sorted(result.sections):
        sec = result.sections[n]
        # 节标题行 (S0 标"页眉/页脚")
        detail_row(n, "页眉/页脚" if n == 0 else "节标题", "", sec.full_title, "",
                   "", "", "不可编辑", "", sec_font, sec_fill)
        sec_rows = [x for x in sec.iter_rows() if x.kind != "section"]
        # 父子链栈: 栈底=节标题. sub/分组标题压栈为新的父级; 字段/说明挂栈顶.
        # 并列 sub (同级序号) 先弹栈到节级再入栈 (8.1/8.2 并列, 8.2 不挂 8.1);
        # 分组标题只吸收其下子项, 并列防护类别 (眼睛/身体防护) 回到分组父级.
        parent_stack = [sec.full_title]
        _GROUP_END_KW = ("眼睛", "身体", "呼吸", "听力", "面部", "头部", "皮肤")
        last_field = ""   # 最近一个非空标签字段 (孤儿节点挂回目标)
        for i, row in enumerate(sec_rows):
            k = resolve_field_key(conn, n, row.label) if row.label else ""
            nxt = sec_rows[i + 1] if i + 1 < len(sec_rows) else None
            is_group = (n != 0 and row.kind == "field" and not (row.value or "").strip()
                        and nxt is not None and nxt.kind in ("field", "sub"))
            if row.kind == "sub":
                # 子标题: 同级 (序号段数相同的 sub) 并列 → 弹栈到节级再入栈
                depth = len([x for x in (row.seq or "").split(".") if x])
                while len(parent_stack) > 1 and depth > 0 and parent_stack[-1].startswith(""):
                    # 弹出前一个 sub/分组标题 (回到节级), 让同级 sub 并列
                    parent_stack = [parent_stack[0]]
                    break
                parent_stack.append(row.label)
                detail_row(n, "子标题", row.seq, row.label, "", "",
                           parent_stack[-2], "不可编辑" if not row.editable else "可编辑",
                           "", sub_font, sub_fill)
            elif is_group:
                parent_stack.append(row.label)
                detail_row(n, "分组标题", row.seq, row.label, "", "",
                           parent_stack[-2], "可编辑", "引导其下子项", sub_font, sub_fill)
            elif row.kind == "field":
                if not (row.label or "").strip():
                    # 孤儿节点 (原文首格空, 如 S5 无标题行) → 挂回最近字段
                    parent = last_field or parent_stack[-1]
                    detail_row(n, "字段", row.seq, "", "", row.value or "",
                               parent, "可编辑" if row.editable else "不可编辑",
                               "无标签（原文首格空），挂接前项", val_font)
                else:
                    # 并列防护类别 (眼睛/身体防护) 在分组标题后 → 回到分组父级
                    if (isinstance(parent_stack[-1], str)
                            and row.label.startswith(_GROUP_END_KW)
                            and len(parent_stack) > 2):
                        parent = parent_stack[-2]
                        parent_stack = parent_stack[:-1]  # 分组标题作用域结束
                    else:
                        parent = parent_stack[-1]
                    last_field = row.label
                    # 空值字段备注: 可编辑 → 待填; 不可编辑 → 固定标题 (如 页眉标题)
                    note = ""
                    if not (row.value or "").strip():
                        note = "空值待填" if row.editable else "固定标题"
                    detail_row(n, "字段", row.seq, row.label, k, row.value or "",
                               parent, "可编辑" if row.editable else "不可编辑",
                               note, val_font)
            else:  # note 说明
                detail_row(n, "说明", row.seq, "", "", row.value or "",
                           parent_stack[-1], "可编辑" if row.editable else "不可编辑",
                           "", note_font)
        if n == 3:
            if sec3 and sec3.component_header:
                detail_row(3, "成分表头", "", "", "", sec3.component_header,
                           "3. 成分/组成资料", "不可编辑", "", val_font)
            for i, comp in enumerate(comps, 1):
                comp_parent = f"成分{i}"
                for part, v in (("名称", comp.name), ("CAS", comp.cas),
                                ("含量", comp.conc)):
                    detail_row(3, "成分", "", f"成分{i}·{part}", "", v,
                               comp_parent, "可编辑", "", val_font)

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 24
    ws2.column_dimensions["E"].width = 20
    ws2.column_dimensions["F"].width = 60
    ws2.column_dimensions["G"].width = 22
    ws2.column_dimensions["H"].width = 9
    ws2.column_dimensions["I"].width = 24

    wb.save(out_path)
    return {"wide_cols": len(headers), "detail_rows": rows_detail,
            "components": len(comps),
            "wide_keys": sum(1 for _n, _k, t in columns if t == "field")}
