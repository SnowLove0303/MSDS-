# -*- coding: utf-8 -*-
"""总库宽表导出: 行=型号, 列=标准字段(纯标准名) + 成分列.

列头 = 'S{节}.{标准字段}' (跨节同名以节前缀去歧义);
一行一个型号; 值聚合同标准字段多行 (换行合并+去重);
一二级类目在第二列, 便于按类目筛选对比.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import catalog


def export_wide(conn, out_path, active_only=True) -> dict:
    data = catalog.materialize_wide(conn, active_only=active_only)
    columns = data["columns"]
    col_keys = data["col_keys"]
    rows = data["rows"]          # (pid, model_name, cell)
    comp_max = data["comp_max"]

    comp_headers = []
    for i in range(1, comp_max + 1):
        for part in ("名称", "CAS", "含量"):
            comp_headers.append(f"S3·成分{i}{part}")

    headers = ["型号", "一级类目"] + columns + comp_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "MSDS总库宽表"
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="2F5597")
    hdr_font = Font(name="Microsoft YaHei", size=9, bold=True, color="FFFFFF")
    val_font = Font(name="Microsoft YaHei", size=9)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = wrap
        c.border = border
    ws.freeze_panes = "B2"

    for ri, (pid, model, cell) in enumerate(rows, start=2):
        p = catalog.get_product(conn, pid)
        cat = (p or {}).get("category_name") or ""
        ws.cell(ri, 1, model).font = Font(name="Microsoft YaHei", size=9, bold=True)
        ws.cell(ri, 1).alignment = wrap
        ws.cell(ri, 1).border = border
        ws.cell(ri, 2, cat).font = val_font
        ws.cell(ri, 2).alignment = wrap
        ws.cell(ri, 2).border = border
        for ci, key in enumerate(col_keys, start=3):
            v = cell.get(key, "")
            c = ws.cell(ri, ci, v)
            c.font = val_font
            c.alignment = wrap
            c.border = border
        base = 3 + len(col_keys)
        for ci in range(base, base + len(comp_headers)):
            idx = ci - base
            i = idx // 3 + 1
            part = ("名称", "CAS", "含量")[idx % 3]
            v = cell.get(("comp", i, part), "")
            c = ws.cell(ri, ci, v)
            c.font = val_font
            c.alignment = wrap
            c.border = border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    for cc in range(3, len(headers) + 2):
        ws.column_dimensions[get_column_letter(cc)].width = 14

    out = Path(out_path)
    wb.save(out)
    return {"models": len(rows), "cols": len(headers)}
